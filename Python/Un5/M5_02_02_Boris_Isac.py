#python 3,10,8
import openpyxl

file = openpyxl.load_workbook("02_Excel_data.xlsx")
print(file.sheetnames)

olimp = file.sheetnames[2]
olimpiadaSheel = file[olimp]

medalhas = [["USA", 46, 12, 5],
            ["China", 38, 20, 7],
            ["UK", 29, 7, 7],
            ["Russia", 22, 10, 9],
            ["South Korea", 13, 3, 2],
            ["Germany", 11, 7, 4]]


for i in medalhas:#Write medalhas in workBook
    olimpiadaSheel.append(i)

olimpiadaSheel.insert_rows(1)#insert empty row in A1

olimpiadaSheel["A1"] = "Pais"
olimpiadaSheel["B1"] = "Ouros"
olimpiadaSheel["C1"] = "Pratas"
olimpiadaSheel["D1"] = "Bronzes"

#print ceel and value
for row in olimpiadaSheel.rows:
    for coll in row:
        print(coll.coordinate , coll.value)

file.save("02_Excel_data.xlsx")
file.close()