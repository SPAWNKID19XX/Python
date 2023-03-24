#Python3.10.8

import openpyxl

file = openpyxl.load_workbook("02_Excel_data.xlsx")#load
listSheet = file.sheetnames#List all sheets
print(listSheet)

if 'Olimpiadas' not in  listSheet: #Create Olimpiadas in file if not existe
    file.create_sheet("Olimpiadas")
else:#else create newSheet
    file.create_sheet("newSheet")

listSheet = file.sheetnames#List all sheets
print(listSheet)
file.save("02_Excel_data.xlsx")
file.close()
