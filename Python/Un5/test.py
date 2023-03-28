import openpyxl, datetime
from openpyxl.chart import AreaChart, Reference

workbook = openpyxl.load_workbook("02_Excel_data.xlsx")

nameSheet = workbook.sheetnames[0]
folha1 = workbook[nameSheet]
'''
print(folha1['A1'].value)
print(folha1.cell(row = 2, column = 3).value)


selectFewCeels = folha1['A1:B2']

for row in selectFewCeels:
    for coll in row:
        print(coll.value)

for line in folha1:
    for coll in line:
        print(coll.coordinate,coll.value)

for row in folha1.rows:
    for collumn in row:
        print(collumn.coordinate, collumn.value)
    print("===fim de linha===")
'''
nameSheet = workbook.sheetnames[1]
folha2 = workbook[nameSheet]
print("Folha 2 = ",folha2.title)

folha2['B6'] = 5
folha2['C6'].value = 5
#folha2['E2'].value = datetime.datetime.now()
folha2['B6'] = 70
folha2['B6'] = '=sum(B6,3)'

#folha2.append(["Boris", 50, 100])

#mySheet = workbook.create_sheet("mySheet")

#workbook['mySheet'].title = "Boris"


workbook.save('02_Excel_data.xlsx')

'''
grafico = AreaChart()
grafico.title = "grafico Excel"
grafico.style = 13
grafico.x_axis.title = "Periodo"
grafico.y_axis.title = "Utilizadores"
periodo = Reference(folha2,min_col = 1, min_row = 2, max_row = 7)
utilizadores = Reference(folha2, min_col=2, min_row=1, max_col=3, max_row = 7)
grafico.add_data(utilizadores, titles_from_data=True)
grafico.set_categories(periodo)
folha2.add_chart(grafico, "A10")
workbook.save('02_Excel_data.xlsx')

for row in folha2.rows:
    for collumn in row:
        print(collumn.coordinate,collumn.value)
'''

import csv
import operator
'''
with open("02_CSV_data.csv", "r", encoding="UTF-8") as myCsv:
    readerCsv = csv.reader(myCsv)
    next(readerCsv)
    for index, row in enumerate(readerCsv):
        print("Linha" + str(index))
        print("Nome: {} Provincia: {} Consumo: {}".format(row[0], row[1], row[2]))

myCsv = csv.reader(open("02_CSV_data.csv", "r", encoding="UTF-8"))
listaFiltrada = sorted(myCsv, key=operator.itemgetter(2), reverse=False)
for i in listaFiltrada:
    print(i)

with open("02_CSV_data.csv", encoding="UTF-8") as myCsv:
    reader = csv.DictReader(myCsv)
    listaOrig = list(reader)
    listOrd = sorted(listaOrig, key=operator.itemgetter("name"), reverse = True)
    for regist in listOrd:
        print(regist)

del reader
myCsv.close()
del myCsv
print("Csv file has been closed")
'''

newDatas = [("ddd", 111),("eee", 222),("fff", 333)]
myCsv = open("03_CSV.csv", "a")
#writer = csv.writer(myCsv)
#writer.writerow(["coluna1" , "coluna2"])
#writer.writerows(newDatas)

#del writer

#writer = csv.writer(myCsv)
#writer.writerow(newDatas)
#myCsv.close()

myCsv = open("03_CSV.csv", "r", encoding="UTF-8")
for row in myCsv:
    print(row)
myCsv.close()
