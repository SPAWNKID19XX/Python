import openpyxl
from openpyxl.styles import Font

file = openpyxl.load_workbook("02_Excel_data.xlsx")

bold = Font(bold=True)#define Font options
olimpSheet = file["Olimpiadas"]#Select Sheet
lastCollumn = len(olimpSheet["1"]) #find last Column
olimpSheet.cell(row=1,column=lastCollumn+1).value = "Total"#save value total to last column
olimpSheet.cell(row=1,column=lastCollumn+1).font = bold

for row in olimpSheet.rows:
    totalCellCoordinate = row[lastCollumn]
    total = 0
    for coll in row:
        if isinstance(coll.value, int):
            total += coll.value
    if total != 0:
        totalCellCoordinate.value = total
    print(totalCellCoordinate.value)

olimpSheet['E8'].value = '=SUM(E2:E7)'
file.save("02_Excel_data.xlsx")
file.close()
